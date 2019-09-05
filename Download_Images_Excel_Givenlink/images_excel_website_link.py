from xlrd import open_workbook
from xlwt import Workbook
import xlsxwriter
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
from io import StringIO
from io import BytesIO
from PIL import Image
import string


# This code was used to download images from a given link in excel sheet
# I used beautifulsoup and html5lib to parse the website for the images
wb = open_workbook('data.xlsx')                                                     
wb_sheet_name="Sheet1"

def f(link,uid):
     r=requests.get(link)
     c=r.content
     soup = BeautifulSoup(c,'html5lib')
     t=soup.findAll("div", {"class": "_2_AcLJ"})
     down(t[0]['style'],uid)
     

    
def down(image,uid):
    try:
        image=image[21:-1]
        print(image)
        r=requests.get(image)
        content=r.content
        img = Image.open(BytesIO(content))
        name1=uid+".jpg"
        name3="img1/"+name1
        img.save(name3)
    except Exception as e:
        print(e)
        pass

for s in wb.sheets():
	if(s.name==wb_sheet_name):
		for r in range(1,1000):
                    try:
                            link  = (s.cell(r,2).value)
                            unique_id = (s.cell(r,0).value)
                            print(unique_id)
                            f(link,unique_id)
                    except Exception as e:
                    	print(e)
                    	pass
